#!/usr/bin/python3

from openpyxl import load_workbook

wb = load_workbook(filename = 'student.xlsx')
ws = wb.active

l = [] * 0

for row in range(2, ws.max_row + 1):
	mid = float(ws['C' + str(row)].value)
	final = float(ws['D' + str(row)].value)
	homework = float(ws['E' + str(row)].value)
	attend = float(ws['F' + str(row)].value)
	total = mid * 0.3 + final * 0.35 + homework * 0.34 + attend

	ws['G' + str(row)] = total

	l.append((total, row))

l.sort(reverse = True)
a = int(len(l) * 0.3)
b = int(len(l) * 0.7)
a_plus = a // 2
b_plus = (b + a) // 2

for i, (total, row) in enumerate(l):
	if total < 40:
		ws['H' + str(row)] = 'F'
	elif i < a:
		ws['H' + str(row)] = 'A'
	elif a <= i < b:
		ws['H' + str(row)] = 'B'
	else:
		ws['H' + str(row)] = 'C'

for i in range(a_plus):
	row = l[i][1]
	ws['H' + str(row)] = 'A+'
for i in range(a + 1, b_plus + 1):
	row = l[i][1]
	ws['H' + str(row)] = 'B+'
for i in range(b + 1, len(l) // 2 + 1):
	row = l[i][1]
	ws['H' + str(row)] = 'C+'

wb.save('student.xlsx')
